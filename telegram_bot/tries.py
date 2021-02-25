import openpyxl
book = openpyxl.open("base.xlsx")
base = book.active
def excel_start_settings():
    base['A1'].value='id пользователя'
    base['b1'].value='Логин Steam'
    base['c1'].value='Пароль Steam'
    base['d1'].value ='Статус оплаты'
    base['e1'].value ='Тип баланса S-статический; D-динамический'
    base['f1'].value = 'Ставки после крашей (1-5)'
    base.merge_cells('f1:j1')
    base['k1'].value ='Участие в лесенках (0-нет, 1-КНК, 2-КНКНК)'
    base['l1'].value ='Фикс. ставка после лесенки'
    base['m1'].value ='Несгораемый баланс'
    base['n1'].value ='Выбранный сайт'
excel_start_settings()
def find_user_in_base (rubish):
    for row in range(1,base.max_row+1):
        print(base[row][0].value)
        if str(base[row][0].value) == rubish:
            return row
        elif row == base.max_row:
            return (0)
book.save("base.xlsx")
rub=input()
find_user_in_base(rub)
book.close()