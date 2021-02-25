import csv
import telebot
from telebot import types
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
token='1609896509:AAHFFntL5mCRpbRNwtb9mLwOU7zj7XrIsWo'
admin_id = 830130638
bot = telebot.TeleBot(token, parse_mode="HTML")
hide_reply_keyboard = types.ReplyKeyboardRemove()
login = ''
password = ''
users = {}
book = openpyxl.open("base.xlsx")#открытие файла
base = book.active#открытие рабочего листа файла (по умолчанию - первый)
def find_user_in_base (rubish): #возвращает номер строки в результате поиска пользователя по id
    for row in range(1,base.max_row+1):
        if str(base[row][0].value) == str(rubish):
            return row
        elif row == base.max_row:
            return (0)
def excel_start_settings(): #подписывает первую строку таблицы (названия столбцов)
    base['A1'].value='id пользователя'
    base['b1'].value='Логин Steam'
    base['c1'].value='Пароль Steam'
    base['d1'].value ='Статус оплаты'
    base['e1'].value ='Тип баланса S-статический; D-динамический'
    base['f1'].value = 'Ставки после крашей (1-5)'
    base.merge_cells('f1:j1') #объединение ячеек для крашей (только подпись)
    base['k1'].value ='Участие в лесенках (0-нет, 1-КНК, 2-КНКНК)'
    base['l1'].value ='Фикс. ставка после лесенки'
    base['m1'].value ='Несгораемый баланс'
    base['n1'].value ='Выбранный сайт'
class StopProgramm(Exception):
    pass

class CSGO_BAND:

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--mute-audio")
    chrome_options.add_argument("--incognito")
    chrome_options.add_argument("window-size=1920x1080")

    data = {
        'login': '',
        'password': '',
        'bets': [0, 0, 0, 0, 0],
        'podushka': 0,
        'type': 1,
        'id': '',
        'code': ''
    }

    def __init__(self, login='', password='', bets=[0, 0.05, 0.3, 0.95, 1], podushka=0, type=1, chat_id=''):
        self.data['login'] = login
        self.data['password'] = password
        self.data['bets'] = bets
        self.data['podushka'] = podushka
        self.data['type'] = type
        self.data['id'] = chat_id
        self.driver = webdriver.Chrome(options=CSGO_BAND.chrome_options)

    def stop_prog(self):
        book.close()
        self.driver.quit()
        raise StopProgramm

    def get_crash(self, n):
        try:
            time.sleep(0.5)
            button = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"[class='graph-labels'] :nth-child({n}) > span")))
            return float(button.text[:-1])
        except Exception:
            self.stop_prog()

    def get_code(self):
        save = self.data['code']
        self.data['code'] = ''
        return save

    def pre_autorisation(self):
        try:
            self.driver.get("https://csgo.band/")
            time.sleep(2)
            WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='btn btn--green steam-login']"))).click()
            WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='steamAccountName']"))).send_keys(self.data['login'])
            WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='steamPassword']"))).send_keys(self.data['password'])
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='imageLogin']"))).click()
            try:
                WebDriverWait(self.driver, 7).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='error_display']")))
                bot.send_message(self.data['id'],
                                 "Вы указали неверные данные Steam, либо при попытке входа появляется CAPTCHA\nПроверьте правильность логина и пароля командой /steam_data")
                self.stop_prog()
            except TimeoutException:
                pass
            bot.send_message(self.data['id'], "Введите код Steam Guard")
        except Exception:
            self.stop_prog()

    def autorisation(self, autocrash='1.2'):
        try:
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='twofactorcode_entry']"))).send_keys(self.get_code())
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='login_twofactorauth_buttonset_entercode'] > div"))).click()
            def enter(self):
                bot.send_message(self.data['id'], "Код неверный, перезапустите бота командой /start")
                # WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='twofactorcode_entry']"))).send_keys(self.get_code())
                # WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='login_twofactorauth_buttonset_incorrectcode'] > div"))).click()
                self.stop_prog()
            while True:
                try:
                    WebDriverWait(self.driver, 7).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='login_twofactorauth_buttonset_incorrectcode']")))
                    enter(self)
                except TimeoutException:
                    break
            # изменение настроек
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='btn-toggler chat-trigger active']"))).click()
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn-sound"))).click()
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".field-input"))).send_keys(Keys.BACK_SPACE)
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".field-input"))).send_keys(autocrash)
        except Exception:
            self.stop_prog()

    def autorisation_check(self):
        save=0
        for i in self.data['bets']:
            if i > 0:
                save=i
                break
        self.click(1)
        bal = self.get_balance()
        self.click(1)
        if save*bal < 0.25:
            bot.send_message(self.data['id'], "На вашем балансе недостаточно средств")
            self.stop_prog()
        else:
            bot.send_message(self.data['id'], "Отлично, на вашем балансе достаточно средств, бот начал работу")
            self.main()

    def click(self, n):
        try:
            for i in range(n):
                time.sleep(0.4)
                WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='checkbox-control__content']"))).click()
        except Exception:
            self.stop_prog()

    def get_balance(self):
        try:
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='cur-u-drops-selected']")))
            sel = self.driver.find_element(By.CSS_SELECTOR, "[class='cur-u-drops-selected__selected']").text[2:]
            total = self.driver.find_element(By.CSS_SELECTOR, "[class='cur-u-drops-selected__total']").text[2:]
            return float(sel)+float(total)
        except Exception:
            self.stop_prog()

    def make_bet(self):
        try:
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='btn-base btn btn--blue make-bet']"))).click()
        except Exception:
            self.stop_prog()

    def change(self, part, bal):
        try:
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='btn btn--blue']"))).click()
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='exchange-filter-maxPrice-field']"))).send_keys(Keys.BACK_SPACE*10)
            time.sleep(1)
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='exchange-filter-maxPrice-field']"))).send_keys(str(round(part*bal - self.data['podushka'], 2)))
            time.sleep(1)
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='withdraw-list__inner'] > button"))).click()
            time.sleep(1)
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='withdraw-footer'] :nth-child(2)"))).click()
        except Exception:
            self.stop_prog()

    def do_afull_process_dynamic(self, num_of_bets):
        self.click(1)
        self.change(self.data['bets'][num_of_bets], self.get_balance())
        self.click(1)
        time.sleep(3)
        self.make_bet()

    def do_afull_process_static(self, num_of_bets, bal):
        self.click(1)
        self.change(self.data['bets'][num_of_bets], bal)
        self.click(1)
        time.sleep(3)
        self.make_bet()

    def wait_crash(self, n):
        crash_n = self.get_crash(n)
        while crash_n >= 1.2:
            crash_n = self.get_crash(n)
        return self.get_crash(1)

    def main(self):

        if self.data['type'] == 1:
            crash_1 = self.get_crash(1)
            while crash_1 < 1.2:
                crash_1 = self.get_crash(1)
            while True:
                crash_1 = self.get_crash(1)
                if crash_1 < 1.2:
                    if self.data['bets'][0] != 0: self.do_afull_process_dynamic(0)
                    crash_1 = self.wait_crash(2)

                    if crash_1 < 1.2:
                        if self.data['bets'][1] != 0: self.do_afull_process_dynamic(1)
                        crash_1 = self.wait_crash(3)

                        if crash_1 < 1.2:
                            if self.data['bets'][2] != 0: self.do_afull_process_dynamic(2)
                            crash_1 = self.wait_crash(4)

                            if crash_1 < 1.2:
                                if self.data['bets'][3] != 0: self.do_afull_process_dynamic(3)
                                crash_1 = self.wait_crash(5)

                                if crash_1 < 1.2:
                                    if self.data['bets'][4] != 0: self.do_afull_process_dynamic(4)
                                    crash_1 = self.wait_crash(6)

                                    if crash_1 < 1.2:
                                        print("Шестерной краш")
                                        break

        if self.data['type'] == 2:
            crash_1 = self.get_crash(1)
            while crash_1 < 1.2:
                crash_1 = self.get_crash(1)
            while True:
                crash_1 = self.get_crash(1)
                if crash_1 < 1.2:
                    self.click(1)
                    bal = self.get_balance()
                    self.click(1)
                    if self.data['bets'][0] != 0: self.do_afull_process_static(0, bal)
                    crash_1 = self.wait_crash(2)

                    if crash_1 < 1.2:
                        if self.data['bets'][1] != 0: self.do_afull_process_static(1, bal)
                        crash_1 = self.wait_crash(3)

                        if crash_1 < 1.2:
                            if self.data['bets'][2] != 0: self.do_afull_process_static(2, bal)
                            crash_1 = self.wait_crash(4)

                            if crash_1 < 1.2:
                                if self.data['bets'][3] != 0: self.do_afull_process_static(3, bal)
                                crash_1 = self.wait_crash(5)

                                if crash_1 < 1.2:
                                    if self.data['bets'][4] != 0: self.do_afull_process_static(4, bal)
                                    crash_1 = self.wait_crash(6)

                                    if crash_1 < 1.2:
                                        print("Шестерной краш")
                                        break

        if self.data['type'] == 3:
            while True:
                crash_1 = self.get_crash(1)
                crash_2 = self.get_crash(2)
                crash_3 = self.get_crash(3)
                crash_4 = self.get_crash(4)
                while crash_1 == self.get_crash(1) or crash_2 == self.get_crash(2) or crash_3 == self.get_crash(3) or crash_4 == self.get_crash(4):
                    time.sleep(0.1)
                self.do_afull_process_dynamic(0)

class CS_FAIL:

    chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument("--headless")
    # chrome_options.add_argument("--disable-dev-shm-usage")
    # chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--mute-audio")
    chrome_options.add_argument("--incognito")
    chrome_options.add_argument("window-size=1920x1080")

    data = {
        'login': '',
        'password': '',
        'bets': [0, 0, 0, 0, 0],
        'podushka': 0,
        'type': 1,
        'id': '',
        'code': ''
    }

    def __init__(self, login='', password='', bets=[0, 0.05, 0.3, 0.95, 1], podushka=0, type=1, chat_id=''):
        self.data['login'] = login
        self.data['password'] = password
        self.data['bets'] = bets
        self.data['podushka'] = podushka
        self.data['type'] = type
        self.data['id'] = chat_id
        self.driver = webdriver.Chrome(options=CS_FAIL.chrome_options)

    def stop_prog(self):
        book.close()
        self.driver.quit()
        raise StopProgramm

    def get_crash(self, n):
        try:
            time.sleep(0.5)
            button = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"[class='xhistory__content ng-star-inserted'] :nth-child({n})")))
            return float(button.text[:-1])
        except Exception:
            self.stop_prog()

    def get_code(self):
        save = self.data['code']
        self.data['code'] = ''
        return save

    def pre_autorisation(self):
        try:
            self.driver.get("https://cs.fail/")
            time.sleep(2)
            WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='btn btn_size_2 btn_success2']"))).click()
            WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='steamAccountName']"))).send_keys(self.data['login'])
            WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='steamPassword']"))).send_keys(self.data['password'])
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='imageLogin']"))).click()
            try:
                WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='error_display']")))
                bot.send_message(self.data['id'],
                                 "Вы указали неверные данные Steam, либо при попытке входа появляется CAPTCHA\nПроверьте правильность логина и пароля командой /steam_data")
                self.stop_prog()
            except TimeoutException:
                pass
            bot.send_message(self.data['id'], "Введите код Steam Guard")
        except Exception:
            self.stop_prog()

    def autorisation(self, autocrash='1.2'):
        try:
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='twofactorcode_entry']"))).send_keys(self.get_code())
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='login_twofactorauth_buttonset_entercode'] > div"))).click()
            def enter(self):
                bot.send_message(self.data['id'], "Код неверный, перезапустите бота командой /start")
                self.stop_prog()
            while True:
                try:
                    WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[id='login_twofactorauth_buttonset_incorrectcode']")))
                    enter(self)
                except TimeoutException:
                    break
            # изменение настроек
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='xbutton xbutton_icon xbutton_size_x34 xbutton_toggle']"))).click()
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='icon volume_icon']"))).click()
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[name='ratio']"))).clear()
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[name='ratio']"))).send_keys(autocrash)
        except Exception:
            self.stop_prog()

    def autorisation_check(self):
        save=0
        for i in self.data['bets']:
            if i > 0:
                save=i
                break
        self.click(1)
        bal = self.get_balance()
        self.click(1)
        if save*bal < 0.25:
            bot.send_message(self.data['id'], "На вашем балансе недостаточно средств")
            self.stop_prog()
        else:
            bot.send_message(self.data['id'], "Отлично, на вашем балансе достаточно средств, бот начал работу")
            self.main()

    def click(self, n):
        try:
            for i in range(n):
                time.sleep(0.4)
                WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='xswitch__content']"))).click()
        except Exception:
            self.stop_prog()

    def get_balance(self):
        try:
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='inventory__head']")))
            sel = self.driver.find_element(By.CSS_SELECTOR, "[class='inventory__selected'] :nth-child(2)").text
            total = self.driver.find_element(By.CSS_SELECTOR, "[class='inventory__balance'] :nth-child(2)").text
            return float(sel)+float(total)
        except Exception:
            self.stop_prog()

    def make_bet(self):
        try:
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='btn btn_block btn_size_create btn_purple ng-star-inserted']"))).click()
        except Exception:
            self.stop_prog()

    def change(self, part, bal):
        try:
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='btn btn_block btn_size_2 btn_default4']"))).click()
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[type='number']"))).send_keys(Keys.BACK_SPACE*10)
            time.sleep(1)
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[type='number']"))).send_keys(str(round(part*bal - self.data['podushka'], 2)))
            time.sleep(1)
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='skins skins_for_window'] > div"))).click()
            time.sleep(1)
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='xbutton xbutton_block xbutton_buy']"))).click()
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[class='shop__close']"))).click()
        except Exception:
            self.stop_prog()

    def do_afull_process_dynamic(self, num_of_bets):
        self.click(1)
        self.change(self.data['bets'][num_of_bets], self.get_balance())
        self.click(2)
        time.sleep(1)
        self.make_bet()

    def do_afull_process_static(self, num_of_bets, bal):
        self.click(1)
        self.change(self.data['bets'][num_of_bets], bal)
        self.click(2)
        time.sleep(1)
        self.make_bet()

    def wait_crash(self, n):
        crash_n = self.get_crash(n)
        while crash_n >= 1.2:
            crash_n = self.get_crash(n)
        return self.get_crash(1)

    def main(self):

        if self.data['type'] == 1:
            crash_1 = self.get_crash(1)
            while crash_1 < 1.2:
                crash_1 = self.get_crash(1)
            while True:
                crash_1 = self.get_crash(1)
                if crash_1 < 1.2:
                    if self.data['bets'][0] != 0: self.do_afull_process_dynamic(0)
                    crash_1 = self.wait_crash(2)

                    if crash_1 < 1.2:
                        if self.data['bets'][1] != 0: self.do_afull_process_dynamic(1)
                        crash_1 = self.wait_crash(3)

                        if crash_1 < 1.2:
                            if self.data['bets'][2] != 0: self.do_afull_process_dynamic(2)
                            crash_1 = self.wait_crash(4)

                            if crash_1 < 1.2:
                                if self.data['bets'][3] != 0: self.do_afull_process_dynamic(3)
                                crash_1 = self.wait_crash(5)

                                if crash_1 < 1.2:
                                    if self.data['bets'][4] != 0: self.do_afull_process_dynamic(4)
                                    crash_1 = self.wait_crash(6)

                                    if crash_1 < 1.2:
                                        print("Шестерной краш")
                                        break

                                    elif self.data['bets'][4] != 0: self.click(1)
                                elif self.data['bets'][3] != 0: self.click(1)
                            elif self.data['bets'][2] != 0: self.click(1)
                        elif self.data['bets'][1] != 0: self.click(1)
                    elif self.data['bets'][0] != 0: self.click(1)

        if self.data['type'] == 2:
            crash_1 = self.get_crash(1)
            while crash_1 < 1.2:
                crash_1 = self.get_crash(1)
            while True:
                crash_1 = self.get_crash(1)
                if crash_1 < 1.2:
                    self.click(1)
                    bal = self.get_balance()
                    self.click(1)
                    if self.data['bets'][0] != 0: self.do_afull_process_static(0, bal)
                    crash_1 = self.wait_crash(2)

                    if crash_1 < 1.2:
                        if self.data['bets'][1] != 0: self.do_afull_process_static(1, bal)
                        crash_1 = self.wait_crash(3)

                        if crash_1 < 1.2:
                            if self.data['bets'][2] != 0: self.do_afull_process_static(2, bal)
                            crash_1 = self.wait_crash(4)

                            if crash_1 < 1.2:
                                if self.data['bets'][3] != 0: self.do_afull_process_static(3, bal)
                                crash_1 = self.wait_crash(5)

                                if crash_1 < 1.2:
                                    if self.data['bets'][4] != 0: self.do_afull_process_static(4, bal)
                                    crash_1 = self.wait_crash(6)

                                    if crash_1 < 1.2:
                                        print("Шестерной краш")
                                        break

                                    elif self.data['bets'][4] != 0: self.click(1)
                                elif self.data['bets'][3] != 0: self.click(1)
                            elif self.data['bets'][2] != 0: self.click(1)
                        elif self.data['bets'][1] != 0: self.click(1)
                    elif self.data['bets'][0] != 0: self.click(1)

        if self.data['type'] == 3:
            i = 0
            while True:
                if i != 0:
                    self.click(1)
                else:
                    i = 1
                crash_1 = self.get_crash(1)
                crash_2 = self.get_crash(2)
                crash_3 = self.get_crash(3)
                crash_4 = self.get_crash(4)
                while crash_1 == self.get_crash(1) or crash_2 == self.get_crash(2) or crash_3 == self.get_crash(3) or crash_4 == self.get_crash(4):
                    time.sleep(0.1)
                self.do_afull_process_dynamic(0)

def set_settings(message):#Запись данных ставок в таблицу
    row=find_user_in_base(message.chat.id)
    markup_reply = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_yes = types.KeyboardButton('Динамический')
    item_no = types.KeyboardButton('Статический')
    markup_reply.add(item_yes, item_no)
    bot.send_message(message.chat.id, "Введите тип баланса (сейчас поясню):\n"
                                      "Ставки будут производиться в процентах баланса (когда от нынешнего баланса вы ставите какую-либо часть), "
                                      "либо они будут зафиксированы (в независимости от размера баланса вы будете ставить одну и ту же сумму, "
                                      "конечно же непривосходящую баланс)\n"
                                      "Динамический баланс позволит вам быстро поднятся, но с ним вы рискуете проиграть почти все\n"
                                      "Статический баланс позволит вам уверенно подниматься, но этот процесс займет больше времени", reply_markup=markup_reply)
    bot.register_next_step_handler(message, set_balance)
def set_balance(message):
    row=find_user_in_base(message.chat.id)
    if message.text.lower()=='динамический':
        base[row][4].value='D'
        bot.send_message(message.chat.id, "Т", reply_markup=hide_reply_keyboard)

    elif message.text.lower() == 'статический':
        base[row][4].value='S'
        bot.send_message(message.chat.id, "Т", reply_markup=hide_reply_keyboard)

    book.save("base.xlsx")
def change_old_user (message):
    us=find_user_in_base(message.chat.id)
    logs = [str(x) for x in message.text.split()]
    base[us][1].value = logs[0]
    base[us][2].value = logs[1]
    book.save("base.xlsx")
def register_new_user(message):
    base[base.max_row+1][0].value=message.chat.id
    logs=[str (x)  for x in message.text.split()]
    base[base.max_row][1].value = logs[0]
    base[base.max_row][2].value = logs[1]
    book.save("base.xlsx")
@bot.message_handler(commands=['start'])
def getting_start(message):
    if (find_user_in_base(message.chat.id) == 0):
        markup_reply = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item_start = types.KeyboardButton('Начать')
        markup_reply.add(item_start)
        bot.send_message(message.chat.id, "Нажми на кнопку, чтобы запустить бота", reply_markup=markup_reply)
    else:
        bot.send_message(message.chat.id, "Просьба не вызывать команду '/start', вы уже зарегестрированы в базе данных")
@bot.message_handler(commands=['steam_data'])
def send_steam_data(message):
    us=find_user_in_base(message.chat.id)
    if us == 0:
        bot.send_message(message.chat.id, "Запрашиваемые данные не обнаружены")
        markup_reply = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item_yes = types.KeyboardButton('Хочу')
        item_no = types.KeyboardButton('Нет')
        markup_reply.add(item_yes, item_no)
        time.sleep(1)
        message = bot.send_message(message.chat.id, "Хотите указать их?", reply_markup=markup_reply)
        bot.register_next_step_handler(message, register_again)
    else:
        bot.send_message(message.chat.id, "Логин Steam: " + str(base[us][1].value)+"\nПароль Steam: "+ str(base[us][2].value))
        markup_reply = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item_yes = types.KeyboardButton('Да')
        item_no = types.KeyboardButton('Нет')
        markup_reply.add(item_yes, item_no)
        time.sleep(1)
        message = bot.send_message(message.chat.id, "Хотите изменить их?", reply_markup=markup_reply)
        bot.register_next_step_handler(message, register_again)

def register_again(message):
    if message.text.lower() == 'да':
        bot.send_message(message.chat.id, "Введите данные Steam, логин и пароль, через пробел", reply_markup=hide_reply_keyboard)
        bot.register_next_step_handler(message, change_old_user)
    elif message.text.lower() == 'хочу':
        bot.send_message(message.chat.id, "Введите данные Steam, логин и пароль, через пробел",reply_markup=hide_reply_keyboard)
        bot.register_next_step_handler(message, register_new_user)
    else:
        bot.send_message(message.chat.id, "Список доступных команд:\n"
                                          "",reply_markup=hide_reply_keyboard)
@bot.message_handler(content_types=['text'])
def start_bot(message):
    if message.text.lower() == "начать":
        bot.send_message(message.chat.id,"Здравствуйте, вы активировали бота, который поможет вам полностью автоматизировать"
                                         " ваши ставки на сайтах csgo.fail, csgo.band!\n"
                                         "Для начала давайте настроим бота (эти параметры можно будет изменить в любой момент).")
        base[base.max_row + 1][0].value = message.chat.id
        set_settings(message)
    elif message.text.lower() == "стоп":
        try:
            users[message.chat.id].stop_prog()
        except StopProgramm:
            try:
                del users[message.chat.id]
                bot.send_message(message.chat.id, "Бот остановлен\nПерезапустите его командой /begin")
            except KeyError:
                pass
    else:
        pass

def change_defolt(message):
    if message.text.lower() == "да":
        pass
    elif message.text.lower() == "нет":
        with open("table.csv", newline='') as table:
            reader = csv.DictReader(table, delimiter=';')
            for row in reader:
                if row['user_id'] == str(message.chat.id):
                    login = row['steam_login']
                    password = row['steam_password']
        bot.send_message(message.chat.id, "Бот запущен\nНапечатайте <b>стоп</b>, чтобы остановить его", reply_markup=hide_reply_keyboard)
        try:
            users[message.chat.id] = CSGO_BAND(login=login, password=password, chat_id=message.chat.id)
            users[message.chat.id].pre_autorisation()
            bot.register_next_step_handler(message, code_add)
        except Exception:
            try:
                del users[message.chat.id]
                bot.send_message(message.chat.id, "Бот остановлен\nПерезапустите его командой /start")
            except KeyError:
                pass
    else:
        pass

def code_add(message):
    try:
        if message.text.lower() == "стоп":
            users[message.chat.id].stop_prog()
        else:
            users[message.chat.id].data['code'] = message.text
            users[message.chat.id].autorisation()
            users[message.chat.id].autorisation_check()
    except Exception:
        try:
            del users[message.chat.id]
            bot.send_message(message.chat.id, "Бот остановлен\nПерезапустите его командой /start")
        except KeyError:
            pass


bot.polling(none_stop=True, interval=0)
